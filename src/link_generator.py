import openpyxl
from selenium import webdriver
import time
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import undetected_chromedriver as uc
import accountInfo

def login(driver):  ### 초기 1회 ###
    # '로그인' 클릭
    driver.find_element(By.CSS_SELECTOR, "a.glue-header__link").click()
    time.sleep(2)
    # 아이디 입력 & '다음' 클릭
    driver.find_element(By.CSS_SELECTOR, "input#identifierId").send_keys(accountInfo.admin_id)
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, "#identifierNext > div > button > span").click()
    time.sleep(2)
    # 비밀번호 입력 & '다음' 클릭
    driver.find_element(By.CSS_SELECTOR, "#password > div.aCsJod.oJeWuf > div > div.Xb9hP > input").send_keys(
        accountInfo.admin_pw)
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, "#passwordNext > div > button > span").click()
    time.sleep(5)

    """### 전화번호 인증창이 나타날 경우
    try:
        phoneNumInput = driver.find_element(By.CSS_SELECTOR, "#phoneNumberId")
        myPhoneNum = input("인증번호를 받을 전화번호를 입력하세요: ")
        phoneNumInput.send_keys(myPhoneNum)
        time.sleep(2)
        ## '다음' 버튼: #idvanyphonecollectNext > div > button > span
        driver.find_element(By.CSS_SELECTOR, "#idvanyphonecollectNext > div > button > span").click()
        time.sleep(1)
        authNumInput = driver.find_element(By.CSS_SELECTOR, "#idvAnyPhonePin")
        myAuthNum = input("문자메시지로 받은 인증번호 6자리를 입력하세요: ")
        authNumInput.send_keys(myAuthNum)
        time.sleep(2)
        ## '다음' 버튼: #idvanyphoneverifyNext > div > button > div.VfPpkd-RLmnJb
        driver.find_element(By.CSS_SELECTOR, "# idvanyphoneverifyNext > div > button > div.VfPpkd-RLmnJb").click()
        print("--Your Phone Number is validated--")
    except: #NoSuchElementException
        print("--Authentification Success--")
    """
    print("--LOGIN COMPLETED--")

def generate_link(driver, linksheet, current_row):
    # '새 회의' 버튼
    driver.find_element(By.CSS_SELECTOR,
                        "#yDmH0d > c-wiz > div > div.S3RDod > div > div.Qcuypc > div.Ez8Iud > div > div.VfPpkd-xl07Ob-XxIAqe-OWXEXe-oYxtQd > div:nth-child(1) > div > button > span").click()
    time.sleep(1)
    # '나중에 진행할 회의 만들기' 버튼
    driver.find_element(By.CSS_SELECTOR,
                        "#yDmH0d > c-wiz > div > div.S3RDod > div > div.Qcuypc > div.Ez8Iud > div > div.VfPpkd-xl07Ob-XxIAqe-OWXEXe-oYxtQd > div:nth-child(2) > div > ul > li:nth-child(2) > span.VfPpkd-StrnGf-rymPhb-b9t22c").click()
    time.sleep(5)
    # 링크 값 가져오기
    try:
        link = driver.find_element(By.CSS_SELECTOR, "#yDmH0d > div.VfPpkd-Sx9Kwc.VfPpkd-Sx9Kwc-OWXEXe-vOE8Lb.cC1eCc.UDxLd.PzCPDd.VKf0Le.u9lF8e.VfPpkd-Sx9Kwc-OWXEXe-FNFY6c > div.VfPpkd-wzTsW > div > div.VfPpkd-cnG4Wd > div > div:nth-child(2) > div > div.NgL38b.CZ8zsc > div.VA2JSc")
        #print("USING OLD CHROME VERSION..")
    except:
        link = driver.find_element(By.CSS_SELECTOR, "#yDmH0d > div.VfPpkd-Sx9Kwc.VfPpkd-Sx9Kwc-OWXEXe-vOE8Lb.cC1eCc.UDxLd.PzCPDd.VKf0Le.u9lF8e.VfPpkd-Sx9Kwc-OWXEXe-FNFY6c > div.VfPpkd-wzTsW > div > div.VfPpkd-cnG4Wd > div > div:nth-child(2) > div > div.Hayy8b")
        #print("USING LATEST CHROME VERSION..")
    print(link.text)

    # 새로운 엑셀 파일 만들기
    link.text
    # 셀에 삽입
    #sheet.cell(row=current_row, column=14).value = link.text
    #mailsheet.cell(row=current_row, column=3).value = link.text
    # 'X' 버튼
    driver.find_element(By.CSS_SELECTOR,
                        "div.u9lF8e div.VfPpkd-oclYLd > button > span > svg").click()
    time.sleep(1)

def set_result_form(mailsheet, tempsheet):
    # 1) mailsheet의 (1,1) ~ (NUM_OF_PEOPLE,3)
    # mailsheet에서 (i,1)값을 가져와 resultsheet에서 찾는다
    # resultsheet에 있는 값만큼
    # NUM_OF_PEOPLE 만큼 반복
    return tempsheet

if __name__=="__main__":
    wb = openpyxl.load_workbook("../data/List.xlsx")
    linksheet = wb.active

    # linksheet의 맨 첫 줄에 있는 숫자 n(링크 개수)을 읽어서, n+2 행부터 쌓는다.
    # n 읽어오기 -> xlsx
    driver = uc.Chrome()
    driver.get('https://meet.google.com/')
    driver.maximize_window()
    time.sleep(2)
    login(driver) #로그인
    for i in range(2, NUM_OF_PEOPLE+2):
        generate_link(driver, mailsheet, i)
    print("END!")
    wb.save("../data/List.xlsx")