import openpyxl
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import undetected_chromedriver as uc
import accountInfo


###### [STEP 1] 수험자 리스트 엑셀 열기 ######
wb = openpyxl.load_workbook("./data/자격시험수험자리스트_1102.xlsx")
originsheet = wb.active


"""
###### [STEP 2] 기본 정보 붙여넣기 ######
"""
if __name__=="__main__":
    def prepare_form(sheet):
        global MAX_ROW
        MAX_ROW = -1

        # 이메일 셀 범위 복사 붙여넣기 (C열-> L열로 복사)
        columns = originsheet.iter_cols(min_col=3, max_col=3)
        for col in columns:
            for cell in col:
                cell_new = originsheet.cell(row=cell.row, column=12, value=cell.value)
                MAX_ROW += 1  # 총 row 개수는 (MAX_ROW-1) 개

        # 참조용 메일 주소 넣기
        originsheet.cell(row=MAX_ROW + 2, column=12).value = "dl_edu_cert@navercorp.com"
        originsheet.cell(row=MAX_ROW + 3, column=12).value = "dl_certification@navercorp.com"

        # 날짜 삽입 (A열-> M열)
        dateValue = (originsheet.cell(row=2, column=1).value) % 10000  ## 1102
        month = (int)(dateValue / 100)
        day = (dateValue % 100)
        dateString = str(month) + "월 " + str(day) + "일"  # 월-일 형태의 스트링으로 변환
        for i in range(0, MAX_ROW + 2):
            originsheet.cell(row=i + 2, column=13).value = dateString  # M열에 삽입
        print("--FORM GENERATED--")

    def login(driver):  ### 초기 1회 ###
        # '로그인' 클릭
        driver.find_element(By.CSS_SELECTOR, "a.glue-header__link").click()
        time.sleep(2)
        # 아이디 입력 & '다음' 클릭
        driver.find_element(By.CSS_SELECTOR, "input#identifierId").send_keys(accountInfo.admin_id)
        time.sleep(2)
        driver.find_element(By.CSS_SELECTOR, "#identifierNext > div > button > span").click()
        time.sleep(2)
        # 비밀번호 입력 & '다음' 클릭
        driver.find_element(By.CSS_SELECTOR, "#password > div.aCsJod.oJeWuf > div > div.Xb9hP > input").send_keys(accountInfo.admin_pw)
        time.sleep(2)
        driver.find_element(By.CSS_SELECTOR, "#passwordNext > div > button > span").click()
        time.sleep(5)
        print("--LOGIN COMPLETED--")
    def generate_link(driver, sheet, current_row):
        driver.find_element(By.CSS_SELECTOR,
                            "#yDmH0d > c-wiz > div > div.S3RDod > div > div.Qcuypc > div.Ez8Iud > div > div.VfPpkd-xl07Ob-XxIAqe-OWXEXe-oYxtQd > div:nth-child(1) > div > button > span").click()
        time.sleep(1)
        driver.find_element(By.CSS_SELECTOR,
                            "#yDmH0d > c-wiz > div > div.S3RDod > div > div.Qcuypc > div.Ez8Iud > div > div.VfPpkd-xl07Ob-XxIAqe-OWXEXe-oYxtQd > div:nth-child(2) > div > ul > li:nth-child(2) > span.VfPpkd-StrnGf-rymPhb-b9t22c").click()
        time.sleep(1)
        # driver.find_element(By.CSS_SELECTOR, "div.VKf0Le.u9lF8e div:nth-child(2) > span > button").click()
        # time.sleep(1)
        link = driver.find_element(By.CSS_SELECTOR,
                                   "#yDmH0d > div.VfPpkd-Sx9Kwc.VfPpkd-Sx9Kwc-OWXEXe-vOE8Lb.cC1eCc.UDxLd.PzCPDd.VKf0Le.u9lF8e.VfPpkd-Sx9Kwc-OWXEXe-FNFY6c > div.VfPpkd-wzTsW > div > div.VfPpkd-cnG4Wd > div > div:nth-child(2) > div > div.NgL38b.CZ8zsc > div.VA2JSc")
        print(link.text)
        sheet.cell(row=current_row, column=14).value = link.text
        driver.find_element(By.CSS_SELECTOR,
                            "div.u9lF8e div.VfPpkd-oclYLd > button > span > svg").click()
        time.sleep(1)

    print("START!")
    prepare_form(originsheet)

    driver = uc.Chrome()
    driver.get('https://meet.google.com/')
    driver.maximize_window()
    time.sleep(2)
    login(driver) #로그인
    for i in range(0, MAX_ROW):
        generate_link(driver, originsheet, i+2)

    print("END!")

###### [STEP 3] Selenium으로 브라우저 원격 제어 ######


        ### 반복 ###
    # '새 회의' 버튼: div.VfPpkd-xl07Ob-XxIAqe-OWXEXe-oYxtQd > div:nth-child(1) > div > button > div.VfPpkd-RLmnJb
    # '나중에 진행할 회의 만들기' 버튼: c-wiz li:nth-child(2) > span.VfPpkd-StrnGf-rymPhb-b9t22c
    # copy 아이콘: div.VKf0Le.u9lF8e div:nth-child(2) > span > button
    # X 아이콘: div.u9lF8e div.VfPpkd-oclYLd > button > span > svg

# step 4. 엑셀에 정보 붙여넣기
#sheet.append([title, genre, audience])

wb.save("./result/temp_1102.xlsx")

"""
# 나중에) 메일 전송용 엑셀 만들기
wb_mail = openpyxl.Workbook()
mailsheet = wb_mail.active
mailsheet.append(["수신자 Email 주소", "date", "link"])  # <--맨 첫 행에 추가
wb_mail.save("./result/MASS_INPUT_FORM_today.xlsx")
"""
